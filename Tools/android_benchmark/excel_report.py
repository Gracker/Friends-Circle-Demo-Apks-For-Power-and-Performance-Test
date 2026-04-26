"""Excel comparison report — Chinese-labelled, multi-sheet.

Sheets (workbook saved as 「Excel对比.xlsx」):
  1. 概览           — run info, devices, total scores
  2. 重点指标       — top active metrics with 中文名 / 含义 / 方向 / values / 差值 / 百分比
  3. 差异对比       — every cross-device delta, sorted by largest gap
  4. 滑动详情       — per-(module, load) FPS / jank breakdown, valid + stuck
  5. 配置差异       — hardware / env diffs only (single-device shows full profile)
  6. 原始指标       — full metric dump (one row per metric × device)
  7. 说明与免责     — advisory metrics with reasons + measurement caveats

Headers and labels are in Chinese; data values stay numeric/string as-is.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .analysis import interpret_findings, metric_dependency_hint
from .reports import (
    SCORE_WEIGHTS,
    compute_category_aggregates,
    compute_config_diffs,
    pct_diff,
)
from .utils import parse_float


# ===== styles =====
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(name="PingFang SC", color="FFFFFF", bold=True)
_SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_SUBHEADER_FONT = Font(name="PingFang SC", bold=True, color="172026")
_BODY_FONT = Font(name="PingFang SC", color="172026")
_ADVISORY_FILL = PatternFill("solid", fgColor="FFF3CD")
_FAILED_FILL = PatternFill("solid", fgColor="F8D7DA")
_GOOD_FILL = PatternFill("solid", fgColor="D4EDDA")
_THIN_BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)


# ===== metric dictionary =====
# Each entry: (中文名, 含义, 大了好/小了好, 类别中文)
_DIR_HIGHER = "大了好"
_DIR_LOWER = "小了好"

METRIC_DICT: Dict[str, Tuple[str, str, str, str]] = {
    # CPU — Native
    "native.cpu_int_single_mops": ("整数单核 Mops/s", "原生整数 ALU 单核吞吐（big 核 pin）", _DIR_HIGHER, "CPU"),
    "native.cpu_int_best_mops": ("整数多核最佳 Mops/s", "原生整数 ALU 多核最高吞吐（1/2/4/8 线程取最佳）", _DIR_HIGHER, "CPU"),
    "native.cpu_fp_single_mflops": ("浮点单核 MFLOPS", "原生浮点 ALU 单核吞吐", _DIR_HIGHER, "CPU"),
    "native.cpu_fp_best_mflops": ("浮点多核最佳 MFLOPS", "原生浮点 ALU 多核最高吞吐", _DIR_HIGHER, "CPU"),
    "native.cpu_int_scaling_pct": ("整数多核扩展效率 %", "best_multi / (cpu_count × single)；衡量并行可扩展性，big.LITTLE 异构通常 40-70% 是正常", _DIR_HIGHER, "CPU"),
    "native.cpu_fp_scaling_pct": ("浮点多核扩展效率 %", "同上，浮点版本", _DIR_HIGHER, "CPU"),
    "cpu.big_int_mops": ("大核单核整数 Mops/s", "pin 到大核（最高 max_freq）跑单线程整数", _DIR_HIGHER, "CPU"),
    "cpu.big_fp_mflops": ("大核单核浮点 MFLOPS", "pin 到大核跑单线程浮点", _DIR_HIGHER, "CPU"),
    "cpu.little_int_mops": ("小核单核整数 Mops/s", "pin 到小核（最低 max_freq）跑单线程整数", _DIR_HIGHER, "CPU"),
    "cpu.int_big_over_little": ("大核/小核 整数比", "big_int_mops / little_int_mops；2-3x 是典型异构 SoC", _DIR_HIGHER, "CPU"),
    "memory.l1_ns": ("L1d 缓存延迟 ns", "≤64 KiB 工作集 pointer-chase 平均延迟", _DIR_LOWER, "Memory"),
    "memory.l2_ns": ("L2 缓存延迟 ns", "128 KiB-1 MiB 工作集", _DIR_LOWER, "Memory"),
    "memory.l3_ns": ("L3 缓存延迟 ns", "1-8 MiB 工作集", _DIR_LOWER, "Memory"),
    "memory.dram_ns": ("DRAM 随机访问延迟 ns", ">>L3 工作集，反映真实主存延迟", _DIR_LOWER, "Memory"),
    "multitask.revisit_to_cold_ratio": ("多任务回切冷启比", "首 app 被踢出后再启动耗时 / 真冷启动；接近 1 表示被淘汰，<0.2 表示仍在内存", _DIR_LOWER, "Memory"),
    "multitask.apps_in_chain": ("多任务链长度", "本次依次启动了多少个 app", _DIR_HIGHER, "Memory"),
    # Geekbench-style
    "gb.aes_round_mib_s": ("AES round 软件吞吐", "AES round 变换软件实现（不走 HW 加速）；衡量字节置换 ALU", _DIR_HIGHER, "CPU"),
    "gb.sha256_mib_s": ("SHA-256 软件吞吐", "SHA-256 软件实现（不走 sha2 扩展指令）", _DIR_HIGHER, "CPU"),
    "gb.fft_mflops": ("FFT GFLOPS（×1000）", "4096 点 Cooley-Tukey radix-2 FFT", _DIR_HIGHER, "CPU"),
    "gb.matmul_mflops": ("矩阵乘 MFLOPS", "256×256 单精度浮点 GEMM；偏 cache 友好", _DIR_HIGHER, "CPU"),
    "gb.nbody_mflops": ("N-body MFLOPS", "256-body Euler 引力模拟", _DIR_HIGHER, "CPU"),
    "gb.sort_mops": ("排序 Mops/s", "qsort 1M 个 uint32_t", _DIR_HIGHER, "CPU"),
    # Memory
    "native.memory_copy_mib_s": ("内存 Copy 带宽", "STREAM Copy: c[i] = a[i]，warmup 已剔除", _DIR_HIGHER, "Memory"),
    "native.memory_scale_mib_s": ("内存 Scale 带宽", "STREAM Scale: b[i] = q*c[i]", _DIR_HIGHER, "Memory"),
    "native.memory_add_mib_s": ("内存 Add 带宽", "STREAM Add: c[i] = a[i]+b[i]", _DIR_HIGHER, "Memory"),
    "native.memory_triad_mib_s": ("内存 Triad 带宽", "STREAM Triad: a[i] = b[i]+q*c[i]，最常引用", _DIR_HIGHER, "Memory"),
    "native.memory_latency_ns": ("内存随机访问延迟 ns", "Pointer-chasing，64 字节 stride，64 MiB 工作集", _DIR_LOWER, "Memory"),
    # Memory pressure
    "memory_pressure.max_resident_mb": ("LMK 触发前最大常驻 MiB", "渐进 mmap 直到被 lmkd kill 之前能保住的内存（绝对值，与 RAM 总量混合）", _DIR_HIGHER, "Memory"),
    "memory_pressure.headroom_pct": ("RAM headroom %", "max_resident_mb / memory_total_mb；不同 RAM 容量设备对比时看这个，不看绝对值", _DIR_HIGHER, "Memory"),
    "memory_pressure.lmkd_kills_during": ("期间 lmkd 击杀进程数", "压测期间 logcat 里观测到的 lmkd kill 事件", _DIR_LOWER, "Memory"),
    # Storage
    "storage.seq_write_mib_s": ("顺序写带宽 MiB/s", "末尾 fsync 的 dd 顺序写", _DIR_HIGHER, "Storage"),
    "storage.random_read_iops_native": ("4K 随机读 IOPS（O_DIRECT）", "pread64 + O_DIRECT；O_DIRECT 不可用时回退到无 O_DIRECT（仍是 4K 随机 pread64）", _DIR_HIGHER, "Storage"),
    "storage.random_write_iops_native": ("4K 随机写 IOPS（O_DIRECT）", "pwrite64 + O_DIRECT，结束 fsync 一次；O_DIRECT 不可用时回退到无 O_DIRECT", _DIR_HIGHER, "Storage"),
    "storage.buffered_read_mib_s": ("缓存命中读 MiB/s", "刚写完同一文件读，几乎全 page cache 命中", _DIR_HIGHER, "Storage"),
    "storage.random_read_iops": ("4K 随机读 IOPS（shell）", "shell while+dd 实现，受 fork 开销限制", _DIR_HIGHER, "Storage"),
    "storage.random_write_iops": ("4K 随机写 IOPS（shell）", "shell while+dd 实现，受 fork 开销限制", _DIR_HIGHER, "Storage"),
    # UX Data
    "ux.regex_mib_per_s": ("文本扫描 MiB/s", "朴素 substring 扫描 ASCII 文本", _DIR_HIGHER, "UX 数据"),
    "ux.json_tokenize_mib_per_s": ("JSON tokenize MiB/s", "JSON-ish 字符流 tokenize 速率", _DIR_HIGHER, "UX 数据"),
    "ux.data_sort_mops": ("数据排序 Mops/s", "500K uint32_t qsort", _DIR_HIGHER, "UX 数据"),
    # UX Image
    "ux.image_mpix_out_per_s": ("图像处理 Mpix/s", "5x5 高斯模糊 + 双线性 2x 缩小 + sepia 色彩转换", _DIR_HIGHER, "UX 图像"),
    # Sustained
    "sustained.stability_ratio": ("稳定性比", "min/max round duration，越接近 1 越稳定", _DIR_HIGHER, "持续性能"),
    "sustained.drift_pct": ("性能漂移 %", "round-1 → round-N 漂移百分比，越小越稳", _DIR_LOWER, "持续性能"),
    # Thermal
    "thermal.max_temp_c": ("压测最大温度 °C", "压测期间最高热区温度", _DIR_LOWER, "热"),
    "thermal.delta_temp_c": ("温度上升 °C", "压测期间最高减初始温度", _DIR_LOWER, "热"),
    "thermal.min_freq_ratio": ("最低频率比", "压测期间最小 cur_freq / max_freq；接近 1 表示未限频", _DIR_HIGHER, "热"),
    # Battery
    "battery.drain_pct": ("电量下降百分比", "采样窗口内 level_pct 下降；越小耗电越低", _DIR_LOWER, "电"),
    "battery.avg_power_mw": ("平均功耗 mW", "|current_now| × |voltage_now| 估算", _DIR_LOWER, "电"),
    # Network
    "network.ping_avg_ms": ("Ping 平均 ms", "ICMP 平均往返延迟", _DIR_LOWER, "网络"),
    "network.ping_jitter_ms": ("Ping 抖动 ms", "ICMP 抖动", _DIR_LOWER, "网络"),
    "network.ping_loss_pct": ("Ping 丢包 %", "ICMP 丢包率", _DIR_LOWER, "网络"),
    # Interaction
    "interaction.dispatch_median_ms": ("点击分发延迟 ms", "MotionEvent 内核 → Activity 分发耗时", _DIR_LOWER, "UI 交互"),
    "interaction.on_draw_median_ms": ("点击到 onDraw ms", "ACTION_DOWN → onDraw 开始；不等于像素可见", _DIR_LOWER, "UI 交互"),
    "interaction.frame_median_ms": ("点击到 frame_callback ms", "ACTION_DOWN → 下一次 Choreographer FrameCallback", _DIR_LOWER, "UI 交互"),
    # Launch
    "launch.best_avg_ms": ("启动平均时长 ms", "am start -W 平均时长；非 root 时是半冷启动", _DIR_LOWER, "UI 启动"),
    # Scroll
    "scroll.avg_fps": ("滑动平均 FPS", "所有滑动用例的 gfxinfo avg_fps 均值（已剔除卡死用例）", _DIR_HIGHER, "UI 滑动"),
    "scroll.p95_frame_ms": ("滑动 P95 帧时间", "95% 帧的渲染时长", _DIR_LOWER, "UI 滑动"),
    "scroll.p99_frame_ms": ("滑动 P99 帧时间", "99% 帧的渲染时长，能反映长尾卡顿", _DIR_LOWER, "UI 滑动"),
    "scroll.avg_jank_pct": ("滑动 Jank 比例 %", "janky 帧 / 总帧 × 100", _DIR_LOWER, "UI 滑动"),
    # GPU
    "gpu.glmark2_score": ("glmark2 总分", "glmark2-es2 综合评分", _DIR_HIGHER, "GPU"),
    # AI
    "ai.best_inference_avg_us": ("AI 最佳推理 µs", "TFLite cpu/xnnpack/gpu/nnapi 中最佳的平均推理延迟", _DIR_LOWER, "AI"),
    "ai.accel_speedup": ("AI 加速比", "CPU avg / 最佳 delegate avg", _DIR_HIGHER, "AI"),
    # Video
    "video.encoder_mb_per_s": ("视频编码 MB/s", "screenrecord 输出大小 / 时间，编码侧 wall-clock", _DIR_HIGHER, "视频"),
    # Install
    "install.duration_sec": ("APK 安装耗时 s", "adb install 整体 wall time", _DIR_LOWER, "其他"),
    # Crypto / Baseline (advisory only)
    "crypto.hash_single_mib_s": ("HW SHA256 单核 MiB/s", "shell sha256sum，走 ARMv8 crypto 扩展；非 CPU 分数", _DIR_HIGHER, "其他（仅参考）"),
    "crypto.hash_best_multi_mib_s": ("HW SHA256 多核 MiB/s", "shell sha256sum 多核；非 CPU 分数", _DIR_HIGHER, "其他（仅参考）"),
    "baseline.kernel_loop_mib_s": ("内核 syscall loop", "dd /dev/zero -> /dev/null；不触 DRAM", _DIR_HIGHER, "其他（仅参考）"),
}

CATEGORY_DICT: Dict[str, str] = {
    "cpu": "CPU 算力", "memory": "内存带宽", "memory_pressure": "内存压力", "storage": "存储",
    "ux_data": "UX 数据处理", "ux_image": "UX 图像处理",
    "scroll": "UI 滑动", "launch": "UI 启动", "interaction": "UI 交互",
    "sustained": "持续性能", "thermal": "热与限频", "battery": "电池",
    "network": "网络", "gpu": "GPU", "ai": "AI/NPU",
    "video": "视频（仅参考）", "install": "安装速度",
    "crypto": "硬件加密（仅参考）", "baseline": "syscall 基线（仅参考）",
}


def _label(metric_key: str, fallback_unit: str = "") -> Tuple[str, str, str, str]:
    """Return (中文名, 含义, 方向, 类别中文); falls back to a generic format."""
    if metric_key in METRIC_DICT:
        return METRIC_DICT[metric_key]
    return (metric_key, "（无中文映射，需补充 METRIC_DICT）", "?", "其他")


def _autosize(ws, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            text = str(cell.value)
            # Roughly count CJK chars as 2 width units.
            width = sum(2 if ord(ch) > 127 else 1 for ch in text)
            if width > max_len:
                max_len = width
        ws.column_dimensions[col_letter].width = min(max_width, max_len + 2)


def _row_header(ws, row: int, headers: Sequence[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = _THIN_BORDER


def _row_data(ws, row: int, values: Sequence[Any], fills: Optional[Sequence[Optional[PatternFill]]] = None) -> None:
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = _BODY_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = _THIN_BORDER
        if fills and col - 1 < len(fills) and fills[col - 1] is not None:
            cell.fill = fills[col - 1]


# ===== sheet builders =====
def _sheet_overview(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("概览")
    ws["A1"] = "Android 综合性能 Benchmark — 概览"
    ws["A1"].font = Font(name="PingFang SC", size=16, bold=True)
    ws.merge_cells("A1:F1")
    info_rows: List[Tuple[str, Any]] = [
        ("开始时间", payload.get("started_at")),
        ("结束时间", payload.get("finished_at")),
        ("Schema 版本", payload.get("schema_version")),
        ("设备数", len(payload.get("devices") or [])),
        ("Preset", (payload.get("config") or {}).get("preset")),
        ("失败设备数", len(payload.get("failures") or [])),
    ]
    row = 3
    for k, v in info_rows:
        ws.cell(row=row, column=1, value=k).font = _SUBHEADER_FONT
        ws.cell(row=row, column=2, value=v).font = _BODY_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="设备总分").font = _SUBHEADER_FONT
    row += 1
    _row_header(ws, row, ["设备", "总分", "类别得分", "备注"])
    row += 1
    scores = payload.get("scores") or {}
    for label, score in scores.items():
        cats = score.get("categories") or {}
        cat_text = "; ".join(
            f"{CATEGORY_DICT.get(k, k)}={v:.1f}" if v is not None else f"{CATEGORY_DICT.get(k, k)}=N/A"
            for k, v in cats.items()
        )
        total = score.get("total")
        _row_data(ws, row, [label, total if total is not None else "N/A", cat_text, score.get("note", "")])
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="设备硬件信息（精简）").font = _SUBHEADER_FONT
    row += 1
    _row_header(ws, row, ["设备", "型号", "SoC", "Android", "内存 MiB", "屏幕", "刷新率", "前台 App"])
    row += 1
    for d in payload.get("devices") or []:
        info = d.get("device_info") or {}
        _row_data(ws, row, [
            d.get("device_label"),
            info.get("model"),
            info.get("soc_model"),
            f"{info.get('android_release')} (sdk {info.get('sdk')})",
            info.get("memory_total_mb"),
            info.get("screen_size"),
            f"peak={info.get('peak_refresh_rate')} min={info.get('min_refresh_rate')}",
            info.get("foreground_package"),
        ])
        row += 1
    _autosize(ws)


def _sheet_key_metrics(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("重点指标")
    devices = payload.get("devices") or []
    labels = [d["device_label"] for d in devices]
    headers = ["类别", "指标（中文）", "含义", "方向", "单位"] + labels
    if len(devices) >= 2:
        headers += ["最佳设备", "最差设备", "差值", "百分比差异 (越大说明差距越大)"]
    headers.append("状态")
    _row_header(ws, 1, headers)

    by_metric: Dict[str, Dict[str, Any]] = {}
    for d in devices:
        for m in d.get("metrics") or []:
            key = m["key"]
            entry = by_metric.setdefault(key, {"meta": m, "values": {}})
            entry["values"][d["device_label"]] = m

    sorted_keys = sorted(by_metric.keys(), key=lambda k: (
        list(METRIC_DICT.keys()).index(k) if k in METRIC_DICT else 9999, k
    ))

    row = 2
    for key in sorted_keys:
        entry = by_metric[key]
        meta = entry["meta"]
        zh_name, zh_explain, zh_dir, zh_cat = _label(key)
        unit = meta.get("unit") or ""
        per_device_vals = [parse_float(entry["values"].get(label, {}).get("value")) for label in labels]
        statuses = [entry["values"].get(label, {}).get("status", "—") for label in labels]
        any_advisory = any(s == "advisory" for s in statuses)
        rowdata: List[Any] = [zh_cat, zh_name, zh_explain, zh_dir, unit]
        for v in per_device_vals:
            rowdata.append(v if v is not None else "")
        if len(devices) >= 2:
            present = [(label, val) for label, val in zip(labels, per_device_vals) if val is not None]
            if len(present) >= 2:
                direction = meta.get("direction", "higher")
                if direction == "lower":
                    best = min(present, key=lambda x: x[1])
                    worst = max(present, key=lambda x: x[1])
                else:
                    best = max(present, key=lambda x: x[1])
                    worst = min(present, key=lambda x: x[1])
                delta = abs(best[1] - worst[1])
                pct = pct_diff(best[1], worst[1], direction)
                rowdata += [best[0], worst[0], round(delta, 4), pct]
            else:
                rowdata += ["", "", "", ""]
        rowdata.append("仅参考" if any_advisory else "正式")

        fills: List[Optional[PatternFill]] = [None] * len(rowdata)
        if any_advisory:
            fills[-1] = _ADVISORY_FILL
        _row_data(ws, row, rowdata, fills)
        row += 1
    _autosize(ws)
    ws.freeze_panes = "F2"


def _sheet_diff(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("差异对比")
    devices = payload.get("devices") or []
    if len(devices) < 2:
        ws["A1"] = "单设备运行 — 无跨设备差异可对比。"
        return
    labels = [d["device_label"] for d in devices]
    headers = ["排名", "类别", "指标（中文）", "方向", "单位", "百分比差异 (绝对值)"] + labels + ["最佳", "最差", "差值"]
    _row_header(ws, 1, headers)

    rows_to_render: List[Tuple[float, List[Any]]] = []
    by_metric: Dict[str, Dict[str, Any]] = {}
    for d in devices:
        for m in d.get("metrics") or []:
            if not m.get("valid_for_score"):
                continue
            entry = by_metric.setdefault(m["key"], {"meta": m, "values": {}})
            entry["values"][d["device_label"]] = parse_float(m["value"])

    for key, entry in by_metric.items():
        meta = entry["meta"]
        zh_name, _, zh_dir, zh_cat = _label(key)
        present = [(label, v) for label, v in entry["values"].items() if v is not None]
        if len(present) < 2:
            continue
        direction = meta.get("direction", "higher")
        if direction == "lower":
            best = min(present, key=lambda x: x[1])
            worst = max(present, key=lambda x: x[1])
        else:
            best = max(present, key=lambda x: x[1])
            worst = min(present, key=lambda x: x[1])
        pct = pct_diff(best[1], worst[1], direction) or 0
        delta = abs(best[1] - worst[1])
        per_device = [entry["values"].get(label) for label in labels]
        row_data = [zh_cat, zh_name, zh_dir, meta.get("unit") or "", abs(pct)] + per_device + [best[0], worst[0], round(delta, 4)]
        rows_to_render.append((abs(pct), row_data))

    rows_to_render.sort(key=lambda x: -x[0])
    for rank, (_, data) in enumerate(rows_to_render, 1):
        rowdata = [rank] + data
        # color bigger gaps red
        gap = data[4]  # already abs(pct)
        fill = None
        if isinstance(gap, (int, float)):
            if gap >= 50: fill = _FAILED_FILL
            elif gap >= 20: fill = _ADVISORY_FILL
            elif gap < 5: fill = _GOOD_FILL
        fills = [None] * len(rowdata)
        if fill: fills[5] = fill
        _row_data(ws, rank + 1, rowdata, fills)
    _autosize(ws)
    ws.freeze_panes = "F2"


def _sheet_scroll(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("滑动详情")
    headers = ["设备", "模块", "Load", "FPS", "P95 ms", "P99 ms", "Jank %", "总帧数", "状态", "说明"]
    _row_header(ws, 1, headers)
    row = 2
    for d in payload.get("devices") or []:
        scroll = (d.get("benchmarks") or {}).get("scroll") or {}
        sm = scroll.get("summary") or {}
        for c in sm.get("valid_cases") or []:
            _row_data(ws, row, [
                d["device_label"], c.get("module"), c.get("load"),
                c.get("fps"), c.get("p95_ms"), c.get("p99_ms"),
                c.get("janky_pct"), c.get("frame_count"),
                "正常",
                f"成功捕获 {c.get('frame_count') or 0} 帧" if c.get("frame_count") else "",
            ])
            row += 1
        for c in sm.get("stuck_cases") or []:
            _row_data(ws, row, [
                d["device_label"], c.get("module"), c.get("load"),
                "—", "—", "—", "—", c.get("frame_count") or 0,
                "卡死/无效", c.get("stuck_reason") or "",
            ], fills=[None] * 8 + [_FAILED_FILL, _FAILED_FILL])
            row += 1
    _autosize(ws)
    ws.freeze_panes = "A2"


def _sheet_config_diffs(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("配置差异")
    devices = payload.get("devices") or []
    if len(devices) < 2:
        ws["A1"] = "单设备运行 — 显示该设备完整配置画像（仅作参考）。"
        ws["A1"].font = _SUBHEADER_FONT
        d = devices[0] if devices else {}
        info = d.get("device_info") or {}
        rows = [
            ("型号", info.get("model")),
            ("SoC", info.get("soc_model")),
            ("Android", f"{info.get('android_release')} (sdk {info.get('sdk')})"),
            ("内存", f"{info.get('memory_total_mb')} MiB total / {info.get('memory_available_mb')} MiB available"),
            ("屏幕", info.get("screen_size")),
            ("刷新率", f"peak={info.get('peak_refresh_rate')} min={info.get('min_refresh_rate')}"),
            ("HWUI", info.get("hwui")),
            ("SELinux", info.get("selinux_mode")),
            ("data_fs", info.get("storage", {}).get("data_fs")),
            ("zram", info.get("mm", {}).get("zram")),
            ("MGLRU", info.get("mm", {}).get("mglru_enabled")),
            ("VM tuning", info.get("mm", {}).get("vm_tuning")),
            ("GPU", info.get("gpu_freq")),
            ("Wi-Fi link", info.get("wifi_link")),
            ("vendor power", info.get("vendor_power_mode")),
            ("battery_gate", d.get("battery_gate")),
            ("env applied", (d.get("env") or {}).get("applied")),
        ]
        for i, (k, v) in enumerate(rows, 3):
            ws.cell(row=i, column=1, value=k).font = _SUBHEADER_FONT
            ws.cell(row=i, column=2, value=str(v)).font = _BODY_FONT
        _autosize(ws)
        return
    diffs = compute_config_diffs(devices)
    labels = [d["device_label"] for d in devices]
    headers = ["字段"] + labels + ["说明（含义提示）"]
    _row_header(ws, 1, headers)
    field_explain = {
        "device.soc_model": "SoC 型号（决定整体性能上限）",
        "device.memory_total_mb": "总内存（影响 multitask 与 LMK 行为）",
        "device.android_release": "Android 大版本",
        "device.cpu_count": "CPU 核数",
        "hwui.renderengine_backend": "SurfaceFlinger RenderEngine 后端（skia-gl / skia-vulkan）",
        "hwui.hwui_renderer": "HWUI 渲染器（Skia/SkiaVk）",
        "hwui.hwui_use_vulkan": "HWUI 是否启用 Vulkan",
        "data_fs.fstype": "/data 文件系统类型",
        "data_fs.noatime": "/data 是否禁 atime（影响小文件 IO）",
        "data_fs.discard": "/data 是否启用 TRIM/discard",
        "mm.zram_enabled": "zram 是否启用",
        "mm.zram_comp_algorithm": "zram 压缩算法",
        "mm.mglru_enabled": "MGLRU 是否启用",
        "vm.swappiness": "内核 swappiness（越大越倾向 swap）",
        "vm.dirty_ratio": "脏页比例阈值",
        "gpu.governor": "GPU 频率调节器",
        "wifi.rssi_dbm": "Wi-Fi 信号强度",
        "battery.charging": "运行时是否充电（影响电池/限频判断）",
        "battery.level_pct": "电量",
        "env.animation_scale": "运行期间应用的动画缩放（默认 1.0）",
        "env.screen_brightness": "运行期间应用的固定亮度",
        "env.zen_mode": "运行期间是否启用勿扰",
    }
    for r_idx, row in enumerate(diffs, 2):
        cells = [row.get("field")] + [row.get(label, "") for label in labels] + [field_explain.get(row.get("field"), "")]
        _row_data(ws, r_idx, cells)
    _autosize(ws)
    ws.freeze_panes = "B2"


def _sheet_raw(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("原始指标")
    headers = ["设备", "类别", "指标 key", "中文名", "含义", "方向", "单位", "值", "状态", "advisory_reason"]
    _row_header(ws, 1, headers)
    row = 2
    for d in payload.get("devices") or []:
        for m in d.get("metrics") or []:
            zh_name, zh_explain, zh_dir, _ = _label(m["key"])
            fills = [None] * len(headers)
            if m.get("status") == "advisory":
                fills[8] = _ADVISORY_FILL
            _row_data(ws, row, [
                d["device_label"], CATEGORY_DICT.get(m.get("category"), m.get("category")),
                m["key"], zh_name, zh_explain, zh_dir, m.get("unit"),
                m.get("value"), m.get("status"), m.get("advisory_reason") or "",
            ], fills=fills)
            row += 1
    _autosize(ws)
    ws.freeze_panes = "A2"


def _sheet_caveats(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("说明与免责")
    ws["A1"] = "测量说明 / 仅参考指标 / 已知偏差"
    ws["A1"].font = Font(name="PingFang SC", size=14, bold=True)
    ws.merge_cells("A1:E1")
    notes = [
        "1. 本工具旨在提供可重复、跨设备可对比的性能画像。某些指标天然受设备配置或环境影响，标记为「仅参考」。",
        "2. CPU 分数完全由 native int/fp + Geekbench 风格软件 kernel 计算，不计入 shell sha256（HW 加速）。",
        "3. 内存分数使用 native STREAM-like Triad / Copy / Scale / Add，已含 warmup 剔除与 mlock。",
        "4. 存储分数以 native pread64+O_DIRECT 真随机 IOPS 为准；shell while+dd 版本受 fork 开销限制，仅参考。",
        "5. 滑动指标已自动剔除 frame_count=0 的「卡死」用例（heavy 负载下 WebView 等模块可能 swipe 不响应），列在「滑动详情」。",
        "6. 启动指标在非 root 设备上是「半冷启动」（无 drop_caches/compile-reset），与 root 后差异较大。",
        "7. 点击响应中 onDraw 表示 draw command 开始，不等于像素可见；真值需要 SF present_fence 关联（本工具暂不实现）。",
        "8. 充电中跑 battery stage 不是有效信号，自动标 advisory。",
        "9. 多设备并行时 thermal cross-talk + USB 共享带宽会扰动结果，建议要么物理隔离要么改顺序模式。",
        "10. video stage 用 screenrecord 编码侧 wall-clock，不是真正的解码 fps，仅作 codec 可用性探针。",
    ]
    for i, n in enumerate(notes, 3):
        ws.cell(row=i, column=1, value=n).font = _BODY_FONT
        ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 90

    row = len(notes) + 5
    ws.cell(row=row, column=1, value="本次运行中的 advisory 指标清单").font = _SUBHEADER_FONT
    row += 1
    _row_header(ws, row, ["设备", "指标", "值", "单位", "advisory_reason"])
    row += 1
    for d in payload.get("devices") or []:
        for m in d.get("metrics") or []:
            if m.get("status") == "advisory":
                zh_name, _, _, _ = _label(m["key"])
                _row_data(ws, row, [
                    d["device_label"], zh_name,
                    m.get("value"), m.get("unit"), m.get("advisory_reason") or "",
                ], fills=[None, None, _ADVISORY_FILL, None, None])
                row += 1
    _autosize(ws)


_SEVERITY_LABELS = {3: "严重", 2: "关注", 1: "提示", 0: "信息"}
_SEVERITY_FILLS = {3: _FAILED_FILL, 2: _ADVISORY_FILL, 1: _ADVISORY_FILL, 0: None}


def _sheet_findings(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("数据解读")
    ws["A1"] = "自动解读 / 跨指标依赖分析"
    ws["A1"].font = Font(name="PingFang SC", size=14, bold=True)
    ws.merge_cells("A1:F1")
    ws["A2"] = ("每条解读包含：标题、严重度、详细说明、相关 metric、可能影响的 device_info/env 参数。"
                "用来快速找到「这个数字慢/快是什么导致的」。")
    ws["A2"].font = _BODY_FONT
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:F2")
    findings = interpret_findings(payload)
    _row_header(ws, 4, ["#", "严重度", "标题", "详细说明", "相关 metric", "相关参数"])
    if not findings:
        _row_data(ws, 5, ["-", "-", "本次没有触发解读规则。", "", "", ""])
    else:
        for i, f in enumerate(findings, 5):
            sev = _SEVERITY_LABELS.get(f["severity"], "?")
            cells = [
                i - 4,
                sev,
                f["title"],
                f["detail"],
                ", ".join(f.get("related_metrics") or []),
                ", ".join(f.get("related_params") or []),
            ]
            fills = [None] * len(cells)
            fills[1] = _SEVERITY_FILLS.get(f["severity"])
            _row_data(ws, i, cells, fills)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 80
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["F"].width = 36
    ws.freeze_panes = "A5"


def generate_excel_report(run_dir: Path, payload: Dict[str, Any]) -> Path:
    wb = Workbook()
    if wb.active.title == "Sheet":
        wb.remove(wb.active)
    _sheet_overview(wb, payload)
    _sheet_findings(wb, payload)
    _sheet_key_metrics(wb, payload)
    _sheet_diff(wb, payload)
    _sheet_scroll(wb, payload)
    _sheet_config_diffs(wb, payload)
    _sheet_raw(wb, payload)
    _sheet_caveats(wb, payload)
    path = run_dir / "Excel对比.xlsx"
    wb.save(path)
    return path
